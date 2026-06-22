#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Petra品牌中国区 360度反馈在线调查系统 - 后端 v2.0
Flask + SQLite
新增：子管理员 + 多语言支持 + 管理员作为被评估者
"""

import os, json, hashlib, hmac, secrets, uuid, csv, io, re
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, send_file, g, make_response
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__, static_folder='static', static_url_path='/static')
app.secret_key = os.environ.get('SECRET_KEY') or secrets.token_hex(32)
CORS(app, supports_credentials=True)

# ========== Token signing (HMAC) ==========
def _sign_token(prefix, uid):
    """Sign a user ID with HMAC-SHA256 to prevent token forgery."""
    key = app.secret_key.encode()
    msg = f"{prefix}:{uid}".encode()
    sig = hmac.new(key, msg, 'sha256').hexdigest()[:16]
    return f"{prefix}:{uid}:{sig}"

def _verify_token(token, expected_prefix):
    """Verify HMAC signature on a signed token. Returns the user ID or None."""
    parts = token.split(':', 2)
    if len(parts) != 3 or parts[0] != expected_prefix:
        return None
    prefix, uid_str, sig = parts
    if _sign_token(prefix, uid_str) == token:
        try:
            return int(uid_str)
        except (ValueError, TypeError):
            return None
    return None

DB_DIR = os.environ.get('DB_DIR', os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.environ.get('DB_PATH') or os.path.join(DB_DIR, 'feedback.db')
os.makedirs(os.path.dirname(DB_PATH) if os.path.dirname(DB_PATH) else DB_DIR, exist_ok=True)

# Track DB creation to detect cold-start re-seeding
_RECOVERY_INFO = None

# ========== Database ==========
import sqlite3

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db

@app.teardown_appcontext
def close_db(e):
    db = g.pop('db', None)
    if db: db.close()

def init_db():
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    db.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            en_name TEXT UNIQUE NOT NULL,
            ch_name TEXT,
            password_hash TEXT NOT NULL,
            department TEXT,
            position TEXT,
            manager_en TEXT,
            manager_ch TEXT,
            role TEXT DEFAULT 'employee',
            admin_level TEXT DEFAULT '',
            status TEXT DEFAULT 'active',
            can_edit BOOLEAN DEFAULT 0,
            lang TEXT DEFAULT 'zh',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS sub_admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            en_name TEXT UNIQUE NOT NULL,
            ch_name TEXT,
            password_hash TEXT NOT NULL,
            permissions TEXT DEFAULT '{"view_results":true,"manage_employees":true,"export_data":true,"manage_settings":false}',
            status TEXT DEFAULT 'active',
            created_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (created_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS admin_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            operator_id INTEGER NOT NULL,
            operator_name TEXT NOT NULL,
            target_id INTEGER NOT NULL,
            target_name TEXT NOT NULL,
            action TEXT NOT NULL,
            old_value TEXT,
            new_value TEXT,
            detail TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (operator_id) REFERENCES users(id),
            FOREIGN KEY (target_id) REFERENCES users(id)
        );

        CREATE INDEX IF NOT EXISTS idx_admin_logs_created ON admin_logs(created_at DESC);
        CREATE INDEX IF NOT EXISTS idx_admin_logs_operator ON admin_logs(operator_id);
        CREATE INDEX IF NOT EXISTS idx_admin_logs_target ON admin_logs(target_id);

        CREATE TABLE IF NOT EXISTS dim1_peer (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            evaluator_id INTEGER NOT NULL,
            target_name TEXT NOT NULL,
            target_dept TEXT,
            collaboration_project TEXT,
            score_communication INTEGER DEFAULT 0,
            score_professional INTEGER DEFAULT 0,
            score_responsibility INTEGER DEFAULT 0,
            score_teamwork INTEGER DEFAULT 0,
            score_problem_solving INTEGER DEFAULT 0,
            strengths TEXT,
            improvements TEXT,
            submitted BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (evaluator_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS dim2_upward (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            evaluator_id INTEGER NOT NULL,
            manager_name TEXT NOT NULL,
            score_goal_setting INTEGER DEFAULT 0,
            score_communication INTEGER DEFAULT 0,
            score_delegation INTEGER DEFAULT 0,
            score_feedback INTEGER DEFAULT 0,
            score_team_climate INTEGER DEFAULT 0,
            score_fairness INTEGER DEFAULT 0,
            strengths TEXT,
            improvements TEXT,
            suggestions TEXT,
            submitted BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (evaluator_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS dim3_downward (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            evaluator_id INTEGER NOT NULL,
            subordinate_name TEXT NOT NULL,
            subordinate_position TEXT,
            score_quality INTEGER DEFAULT 0,
            score_professional INTEGER DEFAULT 0,
            score_initiative INTEGER DEFAULT 0,
            score_teamwork INTEGER DEFAULT 0,
            score_problem_solving INTEGER DEFAULT 0,
            score_customer INTEGER DEFAULT 0,
            strengths TEXT,
            improvements TEXT,
            submitted BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (evaluator_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS dim4_leadership (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            anonymous_token TEXT NOT NULL,
            target_name TEXT NOT NULL,
            relationship TEXT,
            score_strategic INTEGER DEFAULT 0,
            score_communication INTEGER DEFAULT 0,
            score_empowerment INTEGER DEFAULT 0,
            score_innovation INTEGER DEFAULT 0,
            score_integrity INTEGER DEFAULT 0,
            score_execution INTEGER DEFAULT 0,
            score_collaboration INTEGER DEFAULT 0,
            score_emotional INTEGER DEFAULT 0,
            feedback_text TEXT,
            q1_org_mgmt TEXT,
            q2_transparency TEXT,
            q3_engagement TEXT,
            submitted BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );
    ''')
    db.commit()

    # Migration: add admin_level column if not exists (for existing databases)
    try:
        db.execute("ALTER TABLE users ADD COLUMN admin_level TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass  # column already exists

    # Migration: add auto_created column to dim3_downward for locked subordinate entries
    try:
        db.execute("ALTER TABLE dim3_downward ADD COLUMN auto_created BOOLEAN DEFAULT 0")
        db.commit()
    except sqlite3.OperationalError:
        pass  # column already exists

    # Migrate: add lang column if not exists
    try:
        db.execute("ALTER TABLE users ADD COLUMN lang TEXT DEFAULT 'zh'")
        db.commit()
    except sqlite3.OperationalError:
        pass  # Column already exists

    # Migration: update passwords to per-user Excel defaults (2026-06-22)
    # This must run after the settings table is created but before seed_users check,
    # so existing Render deployments get the new passwords without DB wipe.
    pw_migrated = db.execute("SELECT value FROM settings WHERE key='pw_migrated_to_excel'").fetchone()
    if not pw_migrated:
        _default_passwords = {
            'Mursal': 'petra2026406', 'Ali': 'petra2026373', 'Rita': 'petra2026212',
            'Maira': 'petra2026444', 'Carey': 'petra2026108', 'Morpheus': 'petra2026386',
            'Katrina': 'petra2026641',
            'Chase': 'petra2026383', 'Holly': 'petra2026449', 'Ian': 'petra2026194',
            'Summer': 'petra2026197', 'Kylie': 'petra2026293', 'Vanessa': 'petra2026102',
            'Linda': 'petra2026462', 'Jun': 'petra2026236', 'Ming': 'petra2026201',
            'Frank': 'petra2026348', 'Jim': 'petra2026100',
            'Suki': 'petra2026284', 'Sheikh': 'petra2026351', 'Neil': 'petra2026425',
            'Chris': 'petra2026129', 'Jemmy': 'petra2026272', 'Jack': 'petra2026362',
            'Catherine': 'petra2026180', 'Sophie': 'petra2026452',
            'Jocelyn': 'petra2026281', 'Lola': 'petra2026168', 'Molly': 'petra2026462',
        }
        _updated = 0
        for _name, _pw in _default_passwords.items():
            db.execute(
                "UPDATE users SET password_hash=? WHERE en_name=? AND status='active'",
                (generate_password_hash(_pw), _name)
            )
            _updated += 1
        db.execute("INSERT OR IGNORE INTO settings VALUES ('pw_migrated_to_excel', '1')")
        db.commit()
        print(f"[Migration] Updated {_updated} users to per-user default passwords (Excel).")

    # Check if users exist
    count = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if count == 0:
        seed_users(db)

    # Seed settings
    db.execute("INSERT OR IGNORE INTO settings VALUES ('submission_open', '1')")
    db.execute("INSERT OR IGNORE INTO settings VALUES ('cycle_name', '2026H1')")
    db.commit()
    db.close()

def seed_users(db):
    """Initialize database with default employees. Each gets a unique random password."""
    global _RECOVERY_INFO
    employees = [
        # id, en_name, ch_name, department, position, manager_en, role, admin_level
        (1, 'Mursal', 'Mursal Khedri', 'Top Management', '管理层', '', 'admin', 'admin'),
        (2, 'Ali', 'Ali Kaaba', 'Top Management', '管理层', '', 'admin', 'admin'),
        (3, 'Rita', '施利静', 'Top Management', '管理层', '', 'admin', 'admin'),
        (4, 'Maira', 'Maira Mumtaz', 'Financial Management', '财务经理', 'Ali', 'employee', ''),
        (5, 'Carey', '郭梦静', 'Financial Management', '财务主管', 'Rita', 'employee', ''),
        (6, 'Morpheus', '邱燕琳', 'People Management', '人力资源经理', 'Mursal', 'admin', 'super_admin'),
        (7, 'Katrina', '杨雪', 'Supply Chain Management', '产品项目经理', 'Ali', 'employee', ''),
        (8, 'Chase', '黎俊杰', 'Supply Chain Management', '采购经理', 'Ali', 'employee', ''),
        (9, 'Holly', '黄雅欣', 'Supply Chain Management', '产品开发专员', 'Rita', 'employee', ''),
        (10, 'Ian', '王寒', 'Supply Chain Management', '采购跟单专员', 'Rita', 'employee', ''),
        (11, 'Summer', '张萍', 'Supply Chain Management', '采购跟单专员', 'Rita', 'employee', ''),
        (12, 'Kylie', '张影影', 'Supply Chain Management', '供应链专员', 'Rita', 'employee', ''),
        (13, 'Vanessa', '陈茂', 'Supply Chain Management', '产品专员', 'Rita', 'employee', ''),
        (14, 'Linda', '谢金光', 'Supply Chain Management', '物流专员', 'Rita', 'employee', ''),
        (15, 'Jun', '穆世俊', 'Supply Chain Management', '操作师', 'Rita', 'employee', ''),
        (16, 'Ming', '薛佳明', 'Supply Chain Management', '操作员', 'Rita', 'employee', ''),
        (17, 'Frank', '潘绍兴', 'Supply Chain Management', '仓管', 'Rita', 'employee', ''),
        (18, 'Jim', '刘金', 'Supply Chain Management', '仓管', 'Rita', 'employee', ''),
        (19, 'Suki', '苏强', 'Creative', '摄影/摄像师', 'Ali', 'employee', ''),
        (20, 'Sheikh', '刘石洪', 'Creative', '3D设计', 'Ali', 'employee', ''),
        (21, 'Neil', '周颖强', 'Creative', '包装设计', 'Ali', 'employee', ''),
        (22, 'Chris', '陈仁福', 'Business Management', '亚马逊运营经理', 'Ali', 'employee', ''),
        (23, 'Jemmy', '姚满杰', 'Business Management', '亚马逊运营', 'Chris', 'employee', ''),
        (24, 'Jack', '余德洋', 'Business Management', '亚马逊运营', 'Chris', 'employee', ''),
        (25, 'Catherine', '赵玲玲', 'Business Management', 'Etsy运营', 'Ali', 'employee', ''),
        (26, 'Sophie', '董小芙', 'Business Management', '商务拓展及销售经理', 'Mursal', 'employee', ''),
        (27, 'Jocelyn', '朱瑾', 'Petra Spark', '业务销售', 'Ali', 'employee', ''),
        (28, 'Lola', '曾庆会', 'Petra Spark', '高级采购经理', 'Ali', 'employee', ''),
        (29, 'Molly', '张莉', 'Petra Jewelry', '业务经理', 'Rita', 'employee', ''),
    ]
    # Per-user default passwords (from PB360员工信息表.xlsx, 2026-06-22)
    # Katrina not in Excel; password preset below
    default_passwords = {
        'Mursal': 'petra2026406', 'Ali': 'petra2026373', 'Rita': 'petra2026212',
        'Maira': 'petra2026444', 'Carey': 'petra2026108', 'Morpheus': 'petra2026386',
        'Katrina': 'petra2026641',
        'Chase': 'petra2026383', 'Holly': 'petra2026449', 'Ian': 'petra2026194',
        'Summer': 'petra2026197', 'Kylie': 'petra2026293', 'Vanessa': 'petra2026102',
        'Linda': 'petra2026462', 'Jun': 'petra2026236', 'Ming': 'petra2026201',
        'Frank': 'petra2026348', 'Jim': 'petra2026100',
        'Suki': 'petra2026284', 'Sheikh': 'petra2026351', 'Neil': 'petra2026425',
        'Chris': 'petra2026129', 'Jemmy': 'petra2026272', 'Jack': 'petra2026362',
        'Catherine': 'petra2026180', 'Sophie': 'petra2026452',
        'Jocelyn': 'petra2026281', 'Lola': 'petra2026168', 'Molly': 'petra2026462',
    }
    env_pw = os.environ.get('INIT_DEFAULT_PASSWORD')
    recovery_seeds = {}
    for e in employees:
        pw = env_pw or default_passwords.get(e[1], secrets.token_urlsafe(10))
        recovery_seeds[e[1]] = pw
        db.execute(
            "INSERT INTO users (id, en_name, ch_name, password_hash, department, position, manager_en, role, admin_level) VALUES (?,?,?,?,?,?,?,?,?)",
            (e[0], e[1], e[2], generate_password_hash(pw), e[3], e[4], e[5], e[6], e[7])
        )
        print(f"[Seed] {e[1]} seeded.")
    db.commit()
    if env_pw:
        print(f"[Seed] All {len(employees)} users seeded with default password from env var.")
    else:
        print("[Seed] ===== EMERGENCY RECOVERY — INITIAL PASSWORDS =====")
        for name, pw in recovery_seeds.items():
            print(f"  {name}: {pw}")
        print("[Seed] ===== SAVE THESE PASSWORDS NOW =====")
    _RECOVERY_INFO = {
        'action': 'db_recreated',
        'time': datetime.now().isoformat(),
        'total_users': len(employees),
        'default_pw_used': bool(env_pw),
        'seeds': {} if env_pw else recovery_seeds
    }

    # Ensure admin_level is set for existing admin users (idempotent)
    db.execute("UPDATE users SET admin_level='super_admin' WHERE en_name='Morpheus' AND role='admin' AND (admin_level='' OR admin_level IS NULL)")
    db.execute("UPDATE users SET admin_level='admin' WHERE en_name IN ('Mursal','Ali','Rita') AND role='admin' AND (admin_level='' OR admin_level IS NULL)")
    db.commit()

# ========== Auth Helpers ==========
def get_auth_token():
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        return auth[7:]
    return request.cookies.get('token', '')

def _is_admin(user):
    """Check if user is admin or sub_admin"""
    return user['role'] in ('admin', 'sub_admin')

def _get_sub_admin_permissions(db, user_id):
    """Get sub-admin permissions dict"""
    if not user_id:
        return None
    row = db.execute("SELECT permissions FROM sub_admins WHERE id=? AND status='active'", (user_id,)).fetchone()
    if row:
        try:
            return json.loads(row['permissions'])
        except:
            return None
    return None

def _resolve_user(token):
    """Resolve HMAC-signed token to user row. Returns (user_dict, user_type) or (None, None).
    Token format: 'u:<user_id>:<hmac_sig>' or 'sa:<sub_admin_id>:<hmac_sig>'
    """
    db = get_db()

    # Sub-admin token: prefix "sa:"
    sid = _verify_token(token, 'sa')
    if sid is not None:
        sub = db.execute("SELECT * FROM sub_admins WHERE id=? AND status='active'", (sid,)).fetchone()
        if sub:
            d = dict(sub)
            d['role'] = 'sub_admin'
            d['department'] = ''
            d['position'] = '子管理员'
            d['manager_en'] = ''
            d['manager_ch'] = ''
            d['can_edit'] = False
            d['lang'] = 'zh'
            return d, 'sub_admin'
        return None, None

    # Regular user token: prefix "u:"
    uid = _verify_token(token, 'u')
    if uid is not None:
        user = db.execute("SELECT * FROM users WHERE id=? AND status='active'", (uid,)).fetchone()
        if user:
            return dict(user), 'user'
    return None, None

def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = get_auth_token()
        if not token:
            return jsonify({'error': '未登录'}), 401
        user, utype = _resolve_user(token)
        if not user:
            return jsonify({'error': '用户不存在或已禁用'}), 401
        g.user = user
        g.user_type = utype
        return f(*args, **kwargs)
    return decorated

def require_admin(f):
    """Require admin or sub_admin role"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = get_auth_token()
        if not token:
            return jsonify({'error': '未登录'}), 401
        user, utype = _resolve_user(token)
        if not user or not _is_admin(user):
            return jsonify({'error': '需要管理员权限'}), 403
        g.user = user
        g.user_type = utype
        return f(*args, **kwargs)
    return decorated

def require_main_admin(f):
    """Require main admin role only"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = get_auth_token()
        if not token:
            return jsonify({'error': '未登录'}), 401
        user, utype = _resolve_user(token)
        if not user or user['role'] != 'admin':
            return jsonify({'error': '仅主管理员可操作'}), 403
        g.user = user
        g.user_type = utype
        return f(*args, **kwargs)
    return decorated

def require_super_admin(f):
    """Require super_admin level — highest privilege tier"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = get_auth_token()
        if not token:
            return jsonify({'error': '未登录'}), 401
        user, utype = _resolve_user(token)
        if not user or user.get('admin_level') != 'super_admin':
            return jsonify({'error': '仅超级管理员可操作'}), 403
        g.user = user
        g.user_type = utype
        return f(*args, **kwargs)
    return decorated

# ========== Init ==========
init_db()

# ========== Static Files ==========
@app.route('/api/health')
def health_check():
    """健康检查端点 — 用于托管平台就绪探测 + 防休眠"""
    return jsonify({'status': 'ok', 'time': datetime.now().isoformat()})

@app.route('/api/recovery-check')
def recovery_check():
    """Public endpoint: check if DB was recently recreated (no auth needed)."""
    if _RECOVERY_INFO:
        return jsonify({
            'warning': True,
            'message': '数据库已被重建（可能是平台冷启动导致）。所有用户密码已重置为新的随机密码。请查看 Render 日志获取初始密码，或使用 INIT_DEFAULT_PASSWORD 环境变量设置统一的默认密码。',
            'message_en': 'Database has been recreated (likely due to platform cold start). All user passwords have been reset to new random values. Check Render logs for initial passwords, or set INIT_DEFAULT_PASSWORD env var for a uniform default.',
            'recreated_at': _RECOVERY_INFO.get('time'),
            'users_affected': _RECOVERY_INFO.get('total_users'),
            'has_default_pw': _RECOVERY_INFO.get('default_pw_used', False)
        })
    return jsonify({'warning': False, 'message': '数据库状态正常'})

@app.route('/api/db-info', methods=['GET'])
@require_auth
def db_info():
    """Return database status. Super admin sees recovery info if DB was recreated."""
    db = get_db()
    user_count = db.execute("SELECT COUNT(*) FROM users WHERE status='active'").fetchone()[0]
    db_file_exists = os.path.exists(DB_PATH)
    db_file_size = os.path.getsize(DB_PATH) if db_file_exists else 0
    resp = {
        'db_path': DB_PATH,
        'db_exists': db_file_exists,
        'db_size_bytes': db_file_size,
        'user_count': user_count
    }
    if g.user.get('admin_level') == 'super_admin' and _RECOVERY_INFO:
        resp['recovery'] = _RECOVERY_INFO
    return jsonify(resp)

@app.route('/')
def index():
    return app.send_static_file('index.html')

# ========== Auth API ==========
@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        if data is None:
            return jsonify({'error': '请发送 JSON 格式请求'}), 400
        en_name = data.get('en_name', '').strip()
        password = data.get('password', '').strip()
        if not en_name or not password:
            return jsonify({'error': '请输入英文名和密码'}), 400
        db = get_db()

        # Try regular users first
        user = db.execute("SELECT * FROM users WHERE en_name=? AND status='active'", (en_name,)).fetchone()
        if user and check_password_hash(user['password_hash'], password):
            return jsonify({
                'token': _sign_token('u', user['id']),
                'user': {
                    'id': user['id'], 'en_name': user['en_name'], 'ch_name': user['ch_name'],
                    'department': user['department'], 'position': user['position'],
                    'manager_en': user['manager_en'], 'manager_ch': user['manager_ch'],
                    'role': user['role'], 'admin_level': user['admin_level'] or '', 'lang': user['lang'] or 'zh'
                }
            })

        # Try sub_admins
        sub = db.execute("SELECT * FROM sub_admins WHERE en_name=? AND status='active'", (en_name,)).fetchone()
        if sub and check_password_hash(sub['password_hash'], password):
            perms = json.loads(sub['permissions']) if sub['permissions'] else {}
            return jsonify({
                'token': _sign_token('sa', sub['id']),
                'user': {
                    'id': sub['id'], 'en_name': sub['en_name'], 'ch_name': sub['ch_name'],
                    'department': '', 'position': '子管理员',
                    'manager_en': '', 'manager_ch': '',
                    'role': 'sub_admin', 'lang': 'zh',
                    'sub_permissions': perms
                }
            })

        return jsonify({'error': '英文名或密码错误'}), 401
    except Exception as e:
        import traceback, sys
        tb = traceback.format_exc()
        print(f'[LOGIN ERROR] {e}\n{tb}', file=sys.stderr, flush=True)
        return jsonify({
            'error': f'服务器内部错误: {str(e)}',
            'debug_trace': tb.split('\n')[-3:] if app.debug else None
        }), 500

@app.route('/api/me', methods=['GET'])
@require_auth
def get_me():
    u = g.user
    db = get_db()
    sub_open = db.execute("SELECT value FROM settings WHERE key='submission_open'").fetchone()
    can_submit = sub_open and sub_open['value'] == '1'

    resp = {
        'id': u['id'], 'en_name': u['en_name'], 'ch_name': u['ch_name'],
        'department': u.get('department', ''), 'position': u.get('position', ''),
        'manager_en': u.get('manager_en', ''), 'manager_ch': u.get('manager_ch', ''),
        'role': u['role'],
        'admin_level': u.get('admin_level', ''),
        'can_edit': bool(u.get('can_edit', False)),
        'can_submit': can_submit,
        'lang': u.get('lang', 'zh'),
        'user_type': g.user_type,
        'has_manager': bool(u.get('manager_en', '')),  # 新增：是否有上级（用于维度二动态必填）
        'has_subordinates': db.execute("SELECT 1 FROM users WHERE manager_en=? AND status='active' AND id != ? LIMIT 1",
                                 (u['en_name'], u['id'])).fetchone() is not None,
    }

    # Sub-admin permissions
    if u['role'] == 'sub_admin':
        perms = _get_sub_admin_permissions(db, u['id'])
        resp['sub_permissions'] = perms

    return jsonify(resp)

@app.route('/api/admin/reset-all-passwords', methods=['POST'])
@require_super_admin
def reset_all_passwords():
    """Reset all users to their per-user default passwords (from PB360 excel)"""
    db = get_db()
    default_passwords = {
        'Mursal': 'petra2026406', 'Ali': 'petra2026373', 'Rita': 'petra2026212',
        'Maira': 'petra2026444', 'Carey': 'petra2026108', 'Morpheus': 'petra2026386',
        'Katrina': 'petra2026641',
        'Chase': 'petra2026383', 'Holly': 'petra2026449', 'Ian': 'petra2026194',
        'Summer': 'petra2026197', 'Kylie': 'petra2026293', 'Vanessa': 'petra2026102',
        'Linda': 'petra2026462', 'Jun': 'petra2026236', 'Ming': 'petra2026201',
        'Frank': 'petra2026348', 'Jim': 'petra2026100',
        'Suki': 'petra2026284', 'Sheikh': 'petra2026351', 'Neil': 'petra2026425',
        'Chris': 'petra2026129', 'Jemmy': 'petra2026272', 'Jack': 'petra2026362',
        'Catherine': 'petra2026180', 'Sophie': 'petra2026452',
        'Jocelyn': 'petra2026281', 'Lola': 'petra2026168', 'Molly': 'petra2026462',
    }
    env_pw = os.environ.get('INIT_DEFAULT_PASSWORD')
    users = db.execute("SELECT id, en_name FROM users WHERE status='active'").fetchall()
    updated = 0
    for u in users:
        pw = env_pw or default_passwords.get(u['en_name'], 'petra2026')
        db.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(pw), u['id']))
        updated += 1
    
    db.commit()
    
    # Record admin log
    operator_id = g.user['id'] if g.user_type == 'user' else None
    if operator_id:
        db.execute(
            "INSERT INTO admin_logs (operator_id, operator_name, target_id, target_name, action, detail) VALUES (?,?,?,?,?,?)",
            (operator_id, g.user['en_name'], 0, 'ALL_USERS', 'reset_all_passwords', f'Reset {updated} users to per-user default passwords')
        )
        db.commit()
    
    return jsonify({
        'success': True,
        'message': f'Reset {updated} users to per-user default passwords',
        'affected_users': updated
    })

@app.route('/api/change_password', methods=['POST'])
@require_auth
def change_password():
    data = request.get_json()
    old_pw = data.get('old_password', '')
    new_pw = data.get('new_password', '')
    if len(new_pw) < 8 or not re.search(r'[A-Za-z]', new_pw) or not re.search(r'[0-9]', new_pw):
        return jsonify({'error': '密码至少8位，需含字母和数字'}), 400
    db = get_db()

    if g.user_type == 'sub_admin':
        if not check_password_hash(g.user['password_hash'], old_pw):
            return jsonify({'error': '原密码错误'}), 400
        db.execute("UPDATE sub_admins SET password_hash=? WHERE id=?", (generate_password_hash(new_pw), g.user['id']))
    else:
        if not check_password_hash(g.user['password_hash'], old_pw):
            return jsonify({'error': '原密码错误'}), 400
        db.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new_pw), g.user['id']))
    db.commit()
    return jsonify({'message': '密码修改成功'})

# ========== Language API ==========
@app.route('/api/lang', methods=['PUT'])
@require_auth
def update_lang():
    data = request.get_json()
    lang = data.get('lang', 'zh')
    if lang not in ('zh', 'en'):
        return jsonify({'error': '不支持的语言'}), 400
    db = get_db()
    if g.user_type == 'sub_admin':
        # Sub-admins don't have lang column; use cookie
        resp = make_response(jsonify({'message': '语言设置已保存', 'lang': lang}))
        resp.set_cookie('lang', lang, max_age=365*24*3600, path='/')
        return resp
    else:
        db.execute("UPDATE users SET lang=? WHERE id=?", (lang, g.user['id']))
        db.commit()
        resp = make_response(jsonify({'message': '语言设置已保存', 'lang': lang}))
        resp.set_cookie('lang', lang, max_age=365*24*3600, path='/')
        return resp

# ========== User List (for forms) ==========
@app.route('/api/users', methods=['GET'])
@require_auth
def list_users():
    db = get_db()
    users = db.execute("SELECT id, en_name, ch_name, department, position, manager_en FROM users WHERE status='active' ORDER BY department, id").fetchall()
    return jsonify([dict(u) for u in users])

@app.route('/api/managers', methods=['GET'])
@require_auth
def list_managers():
    db = get_db()
    mgrs = db.execute("SELECT DISTINCT en_name, ch_name FROM users WHERE en_name IN ('Mursal','Ali','Rita') ORDER BY id").fetchall()
    return jsonify([dict(m) for m in mgrs])

@app.route('/api/my-subordinates', methods=['GET'])
@require_auth
def get_my_subordinates():
    """Return all employees whose manager_en matches the logged-in user's en_name."""
    db = get_db()
    subs = db.execute(
        "SELECT id, en_name, ch_name, department, position FROM users WHERE manager_en=? AND status='active' AND id != ? ORDER BY id",
        (g.user['en_name'], g.user['id'])
    ).fetchall()
    return jsonify([dict(s) for s in subs])

# ========== Dimension 1: Peer Feedback ==========
@app.route('/api/dim1', methods=['GET'])
@require_auth
def get_dim1():
    db = get_db()
    items = db.execute("SELECT * FROM dim1_peer WHERE evaluator_id=? ORDER BY id", (g.user['id'],)).fetchall()
    return jsonify([dict(i) for i in items])

@app.route('/api/dim1', methods=['POST'])
@require_auth
def save_dim1():
    data = request.get_json()
    item_id = data.get('id')
    db = get_db()
    now = datetime.now().isoformat()

    if item_id:
        existing = db.execute("SELECT * FROM dim1_peer WHERE id=? AND evaluator_id=?", (item_id, g.user['id'])).fetchone()
        if not existing:
            return jsonify({'error': '记录不存在'}), 404
        if existing['submitted'] and not g.user.get('can_edit'):
            return jsonify({'error': '已提交，无法修改。如需修改请联系HR。'}), 403
        db.execute('''UPDATE dim1_peer SET target_name=?, target_dept=?, collaboration_project=?,
            score_communication=?, score_professional=?, score_responsibility=?,
            score_teamwork=?, score_problem_solving=?, strengths=?, improvements=?,
            submitted=?, updated_at=? WHERE id=?''',
            (data.get('target_name',''), data.get('target_dept',''), data.get('collaboration_project',''),
             data.get('score_communication', 0), data.get('score_professional', 0),
             data.get('score_responsibility', 0), data.get('score_teamwork', 0),
             data.get('score_problem_solving', 0), data.get('strengths',''),
             data.get('improvements',''), int(data.get('submitted', False)), now, item_id))
        db.commit()
        if int(data.get('submitted', False)):
            _auto_lock_if_all_submitted(db, g.user)
        return jsonify({'id': item_id, 'message': '保存成功'})
    else:
        cur = db.execute('''INSERT INTO dim1_peer (evaluator_id, target_name, target_dept, collaboration_project,
            score_communication, score_professional, score_responsibility, score_teamwork,
            score_problem_solving, strengths, improvements, submitted, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (g.user['id'], data.get('target_name',''), data.get('target_dept',''),
             data.get('collaboration_project',''), data.get('score_communication',0),
             data.get('score_professional',0), data.get('score_responsibility',0),
             data.get('score_teamwork',0), data.get('score_problem_solving',0),
             data.get('strengths',''), data.get('improvements',''),
             int(data.get('submitted', False)), now, now))
        db.commit()
        if int(data.get('submitted', False)):
            _auto_lock_if_all_submitted(db, g.user)
        return jsonify({'id': cur.lastrowid, 'message': '保存成功'})

@app.route('/api/dim1/<int:item_id>', methods=['DELETE'])
@require_auth
def delete_dim1(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM dim1_peer WHERE id=? AND evaluator_id=?", (item_id, g.user['id'])).fetchone()
    if not item:
        return jsonify({'error': '记录不存在'}), 404
    if item['submitted'] and not g.user.get('can_edit'):
        return jsonify({'error': '已提交，无法删除'}), 403
    db.execute("DELETE FROM dim1_peer WHERE id=?", (item_id,))
    db.commit()
    return jsonify({'message': '删除成功'})

# ========== Dimension 2: Upward Feedback (DELETE) ==========
@app.route('/api/dim2/<int:item_id>', methods=['DELETE'])
@require_auth
def delete_dim2(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM dim2_upward WHERE id=? AND evaluator_id=?", (item_id, g.user['id'])).fetchone()
    if not item:
        return jsonify({'error': '记录不存在'}), 404
    if item['submitted'] and not g.user.get('can_edit'):
        return jsonify({'error': '已提交，无法删除'}), 403
    db.execute("DELETE FROM dim2_upward WHERE id=?", (item_id,))
    db.commit()
    return jsonify({'message': '删除成功'})
@app.route('/api/dim2', methods=['GET'])
@require_auth
def get_dim2():
    db = get_db()
    # Only return the single latest record for this user (each user has at most 1 dim2 record)
    item = db.execute("SELECT * FROM dim2_upward WHERE evaluator_id=? ORDER BY id DESC LIMIT 1", (g.user['id'],)).fetchone()
    if item:
        return jsonify([dict(item)])
    return jsonify([])

@app.route('/api/dim2', methods=['POST'])
@require_auth
def save_dim2():
    data = request.get_json()
    item_id = data.get('id')
    db = get_db()
    now = datetime.now().isoformat()
    if item_id:
        existing = db.execute("SELECT * FROM dim2_upward WHERE id=? AND evaluator_id=?", (item_id, g.user['id'])).fetchone()
        if not existing: return jsonify({'error': '记录不存在'}), 404
        if existing['submitted'] and not g.user.get('can_edit'):
            return jsonify({'error': '已提交，无法修改。如需修改请联系HR。'}), 403
        db.execute('''UPDATE dim2_upward SET manager_name=?, score_goal_setting=?, score_communication=?,
            score_delegation=?, score_feedback=?, score_team_climate=?, score_fairness=?,
            strengths=?, improvements=?, suggestions=?, submitted=?, updated_at=? WHERE id=?''',
            (data.get('manager_name',''), data.get('score_goal_setting',0),
             data.get('score_communication',0), data.get('score_delegation',0),
             data.get('score_feedback',0), data.get('score_team_climate',0),
             data.get('score_fairness',0), data.get('strengths',''),
             data.get('improvements',''), data.get('suggestions',''),
             int(data.get('submitted', False)), now, item_id))
        db.commit()
        if int(data.get('submitted', False)):
            _auto_lock_if_all_submitted(db, g.user)
        return jsonify({'id': item_id, 'message': '保存成功'})
    else:
        existing = db.execute("SELECT id FROM dim2_upward WHERE evaluator_id=?", (g.user['id'],)).fetchone()
        if existing:
            db.execute('''UPDATE dim2_upward SET manager_name=?, score_goal_setting=?, score_communication=?,
                score_delegation=?, score_feedback=?, score_team_climate=?, score_fairness=?,
                strengths=?, improvements=?, suggestions=?, submitted=?, updated_at=? WHERE id=?''',
                (data.get('manager_name',''), data.get('score_goal_setting',0),
                 data.get('score_communication',0), data.get('score_delegation',0),
                 data.get('score_feedback',0), data.get('score_team_climate',0),
                 data.get('score_fairness',0), data.get('strengths',''),
                 data.get('improvements',''), data.get('suggestions',''),
                 int(data.get('submitted', False)), now, existing['id']))
            db.commit()
            if int(data.get('submitted', False)):
                _auto_lock_if_all_submitted(db, g.user)
            return jsonify({'id': existing['id'], 'message': '保存成功'})
        else:
            cur = db.execute('''INSERT INTO dim2_upward (evaluator_id, manager_name, score_goal_setting,
                score_communication, score_delegation, score_feedback, score_team_climate,
                score_fairness, strengths, improvements, suggestions, submitted, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (g.user['id'], data.get('manager_name',''), data.get('score_goal_setting',0),
                 data.get('score_communication',0), data.get('score_delegation',0),
                 data.get('score_feedback',0), data.get('score_team_climate',0),
                 data.get('score_fairness',0), data.get('strengths',''),
                 data.get('improvements',''), data.get('suggestions',''),
                 int(data.get('submitted', False)), now, now))
            db.commit()
            if int(data.get('submitted', False)):
                _auto_lock_if_all_submitted(db, g.user)
            return jsonify({'id': cur.lastrowid, 'message': '保存成功'})

# ========== Dimension 3: Downward Feedback ==========
@app.route('/api/dim3', methods=['GET'])
@require_auth
def get_dim3():
    db = get_db()
    # Auto-create entries for subordinates who don't have records yet
    existing_names = db.execute(
        "SELECT subordinate_name FROM dim3_downward WHERE evaluator_id=?", (g.user['id'],)
    ).fetchall()
    existing_set = {r['subordinate_name'] for r in existing_names}
    subordinates = db.execute(
        "SELECT en_name, ch_name, position FROM users WHERE manager_en=? AND status='active' AND id != ?",
        (g.user['en_name'], g.user['id'])
    ).fetchall()
    now = datetime.now().isoformat()
    for sub in subordinates:
        if sub['en_name'] not in existing_set:
            db.execute('''INSERT INTO dim3_downward (evaluator_id, subordinate_name, subordinate_position,
                auto_created, created_at, updated_at) VALUES (?,?,?,1,?,?)''',
                (g.user['id'], sub['en_name'], sub['position'] or '', now, now))
    db.commit()
    items = db.execute("SELECT * FROM dim3_downward WHERE evaluator_id=? ORDER BY id", (g.user['id'],)).fetchall()
    return jsonify([dict(i) for i in items])

@app.route('/api/dim3', methods=['POST'])
@require_auth
def save_dim3():
    data = request.get_json()
    item_id = data.get('id')
    db = get_db()
    now = datetime.now().isoformat()
    if item_id:
        existing = db.execute("SELECT * FROM dim3_downward WHERE id=? AND evaluator_id=?", (item_id, g.user['id'])).fetchone()
        if not existing: return jsonify({'error': '记录不存在'}), 404
        if existing['submitted'] and not g.user.get('can_edit'):
            return jsonify({'error': '已提交，无法修改。如需修改请联系HR。'}), 403
        db.execute('''UPDATE dim3_downward SET subordinate_name=?, subordinate_position=?,
            score_quality=?, score_professional=?, score_initiative=?, score_teamwork=?,
            score_problem_solving=?, score_customer=?, strengths=?, improvements=?,
            submitted=?, updated_at=? WHERE id=?''',
            (data.get('subordinate_name',''), data.get('subordinate_position',''),
             data.get('score_quality',0), data.get('score_professional',0),
             data.get('score_initiative',0), data.get('score_teamwork',0),
             data.get('score_problem_solving',0), data.get('score_customer',0),
             data.get('strengths',''), data.get('improvements',''),
             int(data.get('submitted', False)), now, item_id))
        db.commit()
        if int(data.get('submitted', False)):
            _auto_lock_if_all_submitted(db, g.user)
        return jsonify({'id': item_id, 'message': '保存成功'})
    else:
        existing = db.execute("SELECT id FROM dim3_downward WHERE evaluator_id=? AND subordinate_name=?", 
                             (g.user['id'], data.get('subordinate_name',''))).fetchone()
        if existing:
            db.execute('''UPDATE dim3_downward SET subordinate_name=?, subordinate_position=?,
                score_quality=?, score_professional=?, score_initiative=?, score_teamwork=?,
                score_problem_solving=?, score_customer=?, strengths=?, improvements=?,
                submitted=?, updated_at=? WHERE id=?''',
                (data.get('subordinate_name',''), data.get('subordinate_position',''),
                 data.get('score_quality',0), data.get('score_professional',0),
                 data.get('score_initiative',0), data.get('score_teamwork',0),
                 data.get('score_problem_solving',0), data.get('score_customer',0),
                 data.get('strengths',''), data.get('improvements',''),
                 int(data.get('submitted', False)), now, existing['id']))
            db.commit()
            if int(data.get('submitted', False)):
                _auto_lock_if_all_submitted(db, g.user)
            return jsonify({'id': existing['id'], 'message': '保存成功'})
        else:
            cur = db.execute('''INSERT INTO dim3_downward (evaluator_id, subordinate_name, subordinate_position,
                score_quality, score_professional, score_initiative, score_teamwork,
                score_problem_solving, score_customer, strengths, improvements,
                submitted, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (g.user['id'], data.get('subordinate_name',''), data.get('subordinate_position',''),
                 data.get('score_quality',0), data.get('score_professional',0),
                 data.get('score_initiative',0), data.get('score_teamwork',0),
                 data.get('score_problem_solving',0), data.get('score_customer',0),
                 data.get('strengths',''), data.get('improvements',''),
                 int(data.get('submitted', False)), now, now))
            db.commit()
            if int(data.get('submitted', False)):
                _auto_lock_if_all_submitted(db, g.user)
            return jsonify({'id': cur.lastrowid, 'message': '保存成功'})

@app.route('/api/dim3/<int:item_id>', methods=['DELETE'])
@require_auth
def delete_dim3(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM dim3_downward WHERE id=? AND evaluator_id=?", (item_id, g.user['id'])).fetchone()
    if not item:
        return jsonify({'error': '记录不存在'}), 404
    if item['submitted'] and not g.user.get('can_edit'):
        return jsonify({'error': '已提交，无法删除'}), 403
    db.execute("DELETE FROM dim3_downward WHERE id=?", (item_id,))
    db.commit()
    return jsonify({'message': '删除成功'})

# ========== Helper: Auto-lock after all mandatory dims submitted ==========
def _auto_lock_if_all_submitted(db, user):
    """检查用户是否已完成所有必填维度提交，若是则自动关闭编辑权限"""
    u = db.execute("SELECT * FROM users WHERE id=?", (user['id'],)).fetchone()
    # 必须 can_edit=1 才考虑自动锁闭
    if not u['can_edit']:
        return
    # 检查维度二（仅当有上级时必填）
    has_manager = bool(u.get('manager_en', ''))
    dim2_submitted = True
    if has_manager:
        d2 = db.execute("SELECT 1 FROM dim2_upward WHERE evaluator_id=? AND submitted=1", (user['id'],)).fetchone()
        dim2_submitted = d2 is not None
    # 检查维度三（仅当有下属时必填）
    has_subordinates = db.execute(
        "SELECT 1 FROM users WHERE manager_en=? AND status='active' AND id != ? LIMIT 1",
        (u['en_name'], u['id'])
    ).fetchone() is not None
    dim3_submitted = True
    if has_subordinates:
        d3 = db.execute("SELECT 1 FROM dim3_downward WHERE evaluator_id=? AND submitted=1", (user['id'],)).fetchone()
        dim3_submitted = d3 is not None
    # 所有必填维度已提交 → 自动锁闭
    if dim2_submitted and dim3_submitted:
        db.execute("UPDATE users SET can_edit=0 WHERE id=?", (user['id'],))
        db.commit()

# ========== Dimension 4: Anonymous Leadership ==========
@app.route('/api/dim4/token', methods=['POST'])
@require_auth
def get_dim4_token():
    db = get_db()
    token = request.cookies.get('anon_token', '')
    if token and db.execute("SELECT id FROM dim4_leadership WHERE anonymous_token=?", (token,)).fetchone():
        return jsonify({'token': token})
    token = hashlib.sha256(f"{g.user['id']}-{uuid.uuid4()}".encode()).hexdigest()[:32]
    return jsonify({'token': token})

@app.route('/api/dim4', methods=['GET'])
@require_auth
def get_dim4():
    token = request.args.get('token', '')
    if not token: return jsonify([])
    db = get_db()
    items = db.execute("SELECT * FROM dim4_leadership WHERE anonymous_token=?", (token,)).fetchall()
    return jsonify([dict(i) for i in items])

@app.route('/api/dim4', methods=['POST'])
@require_auth
def save_dim4():
    data = request.get_json()
    token = data.get('token', '')
    if not token:
        return jsonify({'error': '缺少匿名标识'}), 400
    item_id = data.get('id')
    db = get_db()
    now = datetime.now().isoformat()
    if item_id:
        existing = db.execute("SELECT * FROM dim4_leadership WHERE id=? AND anonymous_token=?", (item_id, token)).fetchone()
        if not existing: return jsonify({'error': '记录不存在'}), 404
        if existing['submitted'] and not g.user.get('can_edit'):
            return jsonify({'error': '已提交，无法修改。如需修改请联系HR。'}), 403
        db.execute('''UPDATE dim4_leadership SET target_name=?, relationship=?,
            score_strategic=?, score_communication=?, score_empowerment=?, score_innovation=?,
            score_integrity=?, score_execution=?, score_collaboration=?, score_emotional=?,
            feedback_text=?, q1_org_mgmt=?, q2_transparency=?, q3_engagement=?,
            submitted=?, updated_at=? WHERE id=?''',
            (data.get('target_name',''), data.get('relationship',''),
             data.get('score_strategic',0), data.get('score_communication',0),
             data.get('score_empowerment',0), data.get('score_innovation',0),
             data.get('score_integrity',0), data.get('score_execution',0),
             data.get('score_collaboration',0), data.get('score_emotional',0),
             data.get('feedback_text',''), data.get('q1_org_mgmt',''),
             data.get('q2_transparency',''), data.get('q3_engagement',''),
             int(data.get('submitted', False)), now, item_id))
        db.commit()
        return jsonify({'id': item_id, 'message': '保存成功'})
    else:
        cur = db.execute('''INSERT INTO dim4_leadership (anonymous_token, target_name, relationship,
            score_strategic, score_communication, score_empowerment, score_innovation,
            score_integrity, score_execution, score_collaboration, score_emotional,
            feedback_text, q1_org_mgmt, q2_transparency, q3_engagement, submitted, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (token, data.get('target_name',''), data.get('relationship',''),
             data.get('score_strategic',0), data.get('score_communication',0),
             data.get('score_empowerment',0), data.get('score_innovation',0),
             data.get('score_integrity',0), data.get('score_execution',0),
             data.get('score_collaboration',0), data.get('score_emotional',0),
             data.get('feedback_text',''), data.get('q1_org_mgmt',''),
             data.get('q2_transparency',''), data.get('q3_engagement',''),
             int(data.get('submitted', False)), now, now))
        db.commit()
        return jsonify({'id': cur.lastrowid, 'message': '保存成功'})

# ========== Admin API ==========
@app.route('/api/admin/dashboard', methods=['GET'])
@require_admin
def admin_dashboard():
    db = get_db()
    total_users = db.execute("SELECT COUNT(*) FROM users WHERE status='active'").fetchone()[0]

    stats = {}
    for dim, table, uid_col in [
        ('dim1', 'dim1_peer', 'evaluator_id'),
        ('dim2', 'dim2_upward', 'evaluator_id'),
    ]:
        submitted = db.execute(f"SELECT COUNT(DISTINCT {uid_col}) FROM {table} WHERE submitted=1").fetchone()[0]
        stats[dim] = {'submitted': submitted, 'total': total_users}
    
    # dim3: only count auto_created records (preset subordinates), exclude manual entries
    dim3_submitted = db.execute(
        "SELECT COUNT(DISTINCT evaluator_id) FROM dim3_downward WHERE submitted=1 AND auto_created=1"
    ).fetchone()[0]
    dim3_total = db.execute(
        "SELECT COUNT(DISTINCT u.id) FROM users u WHERE u.status='active' AND EXISTS (SELECT 1 FROM users s WHERE s.manager_en=u.en_name AND s.status='active' AND s.id != u.id)"
    ).fetchone()[0]
    stats['dim3'] = {'submitted': dim3_submitted, 'total': dim3_total}

    dim4_count = db.execute("SELECT COUNT(*) FROM dim4_leadership WHERE submitted=1").fetchone()[0]
    stats['dim4'] = {'submitted': dim4_count, 'total': total_users}

    dept_stats = db.execute('''
        SELECT department, COUNT(*) as cnt FROM users WHERE status='active' GROUP BY department
    ''').fetchall()

    return jsonify({
        'total_users': total_users,
        'submission_stats': stats,
        'department_stats': [dict(d) for d in dept_stats]
    })

@app.route('/api/admin/submissions/<dim>', methods=['GET'])
@require_admin
def admin_submissions(dim):
    db = get_db()
    if dim == 'dim1':
        rows = db.execute('''
            SELECT d.*, u.en_name as evaluator_en, u.ch_name as evaluator_ch, u.department as evaluator_dept
            FROM dim1_peer d JOIN users u ON d.evaluator_id = u.id ORDER BY u.department, d.id
        ''').fetchall()
    elif dim == 'dim2':
        rows = db.execute('''
            SELECT d.*, u.en_name as evaluator_en, u.ch_name as evaluator_ch, u.department as evaluator_dept
            FROM dim2_upward d JOIN users u ON d.evaluator_id = u.id ORDER BY u.department
        ''').fetchall()
    elif dim == 'dim3':
        rows = db.execute('''
            SELECT d.*, u.en_name as evaluator_en, u.ch_name as evaluator_ch, u.department as evaluator_dept
            FROM dim3_downward d JOIN users u ON d.evaluator_id = u.id ORDER BY u.department, d.id
        ''').fetchall()
    elif dim == 'dim4':
        rows = db.execute("SELECT * FROM dim4_leadership ORDER BY submitted DESC, id").fetchall()
    else:
        return jsonify({'error': '无效维度'}), 400
    return jsonify([dict(r) for r in rows])

@app.route('/api/admin/users', methods=['GET'])
@require_admin
def admin_users():
    db = get_db()
    users = db.execute('''
        SELECT u.*,
            (SELECT COUNT(*) FROM dim1_peer WHERE evaluator_id=u.id AND submitted=1) as dim1_done,
            (SELECT COUNT(*) FROM dim2_upward WHERE evaluator_id=u.id AND submitted=1) as dim2_done,
            (SELECT COUNT(*) FROM dim3_downward WHERE evaluator_id=u.id AND submitted=1) as dim3_done
        FROM users u WHERE status='active' ORDER BY u.department, u.id
    ''').fetchall()
    return jsonify([dict(u) for u in users])

@app.route('/api/admin/users/<int:uid>', methods=['PUT'])
@require_admin
def admin_update_user(uid):
    data = request.get_json()
    db = get_db()
    if 'can_edit' in data:
        db.execute("UPDATE users SET can_edit=? WHERE id=?", (int(data['can_edit']), uid))
    if 'password' in data and data['password']:
        db.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(data['password']), uid))
    db.commit()
    return jsonify({'message': '更新成功'})

@app.route('/api/admin/settings', methods=['GET'])
@require_admin
def admin_get_settings():
    db = get_db()
    rows = db.execute("SELECT * FROM settings").fetchall()
    return jsonify({r['key']: r['value'] for r in rows})

@app.route('/api/admin/settings', methods=['PUT'])
@require_admin
def admin_update_settings():
    data = request.get_json()
    db = get_db()
    for k, v in data.items():
        db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)", (k, str(v)))
    db.commit()
    return jsonify({'message': '设置已更新'})

@app.route('/api/admin/unlock/<int:uid>', methods=['POST'])
@require_admin
def admin_unlock_user(uid):
    """Toggle can_edit: if currently closed (0), open it; if open (1), close it.
       When opening, also reset submitted=0 for all dimensions so user can re-submit."""
    db = get_db()
    user = db.execute("SELECT can_edit FROM users WHERE id=?", (uid,)).fetchone()
    if not user:
        return jsonify({'error': '用户不存在'}), 404
    new_val = 0 if user['can_edit'] else 1
    db.execute("UPDATE users SET can_edit=? WHERE id=?", (new_val, uid))
    if new_val == 1:  # opening edit: also reset submitted flags
        for table in ['dim1_peer', 'dim2_upward', 'dim3_downward']:
            db.execute(f"UPDATE {table} SET submitted=0 WHERE evaluator_id=?", (uid,))
    db.commit()
    action = '已开放修改权限' if new_val == 1 else '已关闭修改权限'
    return jsonify({'message': action, 'can_edit': bool(new_val)})

@app.route('/api/admin/reset-password/<int:uid>', methods=['POST'])
@require_super_admin
def admin_reset_password(uid):
    """Super admin emergency password reset for any user. Uses per-user default if no custom pw given."""
    data = request.get_json() or {}
    new_pw = data.get('new_password')
    db = get_db()
    target = db.execute("SELECT * FROM users WHERE id=? AND status='active'", (uid,)).fetchone()
    if not target:
        return jsonify({'error': 'User not found or disabled'}), 404
    if not new_pw:
        default_passwords = {
            'Mursal': 'petra2026406', 'Ali': 'petra2026373', 'Rita': 'petra2026212',
            'Maira': 'petra2026444', 'Carey': 'petra2026108', 'Morpheus': 'petra2026386',
            'Katrina': 'petra2026641',
            'Chase': 'petra2026383', 'Holly': 'petra2026449', 'Ian': 'petra2026194',
            'Summer': 'petra2026197', 'Kylie': 'petra2026293', 'Vanessa': 'petra2026102',
            'Linda': 'petra2026462', 'Jun': 'petra2026236', 'Ming': 'petra2026201',
            'Frank': 'petra2026348', 'Jim': 'petra2026100',
            'Suki': 'petra2026284', 'Sheikh': 'petra2026351', 'Neil': 'petra2026425',
            'Chris': 'petra2026129', 'Jemmy': 'petra2026272', 'Jack': 'petra2026362',
            'Catherine': 'petra2026180', 'Sophie': 'petra2026452',
            'Jocelyn': 'petra2026281', 'Lola': 'petra2026168', 'Molly': 'petra2026462',
        }
        new_pw = default_passwords.get(target['en_name'], 'petra2026')
    if len(new_pw) < 8 or not re.search(r'[A-Za-z]', new_pw) or not re.search(r'[0-9]', new_pw):
        new_pw = new_pw + 'Aa1'  # ensure compliance
    db.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new_pw), uid))
    _log_admin_action(db, g.user, uid, target['en_name'], 'reset_password',
                      '[redacted]', '[redacted]', f"由 {g.user['en_name']} 重置了 {target['en_name']} 的密码")
    return jsonify({
        'message': f'已重置 {target["en_name"]} 的密码',
        'new_password': new_pw,
        'note': '请将新密码安全地告知用户，此密码不会再次显示。'
    })

# ========== Sub-Admin Management (Main Admin Only) ==========
@app.route('/api/admin/sub-admins', methods=['GET'])
@require_main_admin
def list_sub_admins():
    db = get_db()
    subs = db.execute("SELECT * FROM sub_admins ORDER BY id").fetchall()
    result = []
    for s in subs:
        d = dict(s)
        d.pop('password_hash', None)
        d['permissions'] = json.loads(d['permissions']) if d['permissions'] else {}
        result.append(d)
    return jsonify(result)

@app.route('/api/admin/sub-admins', methods=['POST'])
@require_main_admin
def create_sub_admin():
    data = request.get_json()
    en_name = data.get('en_name', '').strip()
    ch_name = data.get('ch_name', '').strip()
    password = data.get('password', '').strip()
    permissions = data.get('permissions', {
        'view_results': True,
        'manage_employees': True,
        'export_data': True,
        'manage_settings': False
    })

    if not en_name or not password:
        return jsonify({'error': '英文名和密码不能为空'}), 400
    if len(password) < 8 or not re.search(r'[A-Za-z]', password) or not re.search(r'[0-9]', password):
        return jsonify({'error': '密码至少8位，需含字母和数字'}), 400

    db = get_db()
    # Check count limit
    count = db.execute("SELECT COUNT(*) FROM sub_admins WHERE status='active'").fetchone()[0]
    if count >= 3:
        return jsonify({'error': '子管理员最多3名'}), 400

    # Check uniqueness
    existing = db.execute("SELECT id FROM sub_admins WHERE en_name=?", (en_name,)).fetchone()
    if existing:
        return jsonify({'error': '该英文名已被使用'}), 400
    # Also check regular users
    existing_user = db.execute("SELECT id FROM users WHERE en_name=?", (en_name,)).fetchone()
    if existing_user:
        return jsonify({'error': '该英文名与已有员工冲突'}), 400

    db.execute('''INSERT INTO sub_admins (en_name, ch_name, password_hash, permissions, created_by)
        VALUES (?,?,?,?,?)''',
        (en_name, ch_name, generate_password_hash(password), json.dumps(permissions), g.user['id']))
    db.commit()
    return jsonify({'message': '子管理员创建成功'})

@app.route('/api/admin/sub-admins/<int:sid>', methods=['PUT'])
@require_main_admin
def update_sub_admin(sid):
    data = request.get_json()
    db = get_db()
    sub = db.execute("SELECT * FROM sub_admins WHERE id=?", (sid,)).fetchone()
    if not sub:
        return jsonify({'error': '子管理员不存在'}), 404

    if 'password' in data and data['password']:
        db.execute("UPDATE sub_admins SET password_hash=? WHERE id=?", (generate_password_hash(data['password']), sid))
    if 'permissions' in data:
        db.execute("UPDATE sub_admins SET permissions=? WHERE id=?", (json.dumps(data['permissions']), sid))
    if 'status' in data:
        db.execute("UPDATE sub_admins SET status=? WHERE id=?", (data['status'], sid))

    db.commit()
    return jsonify({'message': '子管理员已更新'})

@app.route('/api/admin/sub-admins/<int:sid>', methods=['DELETE'])
@require_main_admin
def delete_sub_admin(sid):
    db = get_db()
    db.execute("UPDATE sub_admins SET status='inactive' WHERE id=?", (sid,))
    db.commit()
    return jsonify({'message': '子管理员已禁用'})

# ========== Export ==========
@app.route('/api/admin/export/<dim>', methods=['GET'])
@require_admin
def admin_export(dim):
    fmt = request.args.get('format', 'xlsx')
    db = get_db()

    if fmt == 'csv':
        return export_csv(db, dim)
    else:
        return export_xlsx(db, dim)

def export_xlsx(db, dim):
    wb = openpyxl.Workbook()
    ws = wb.active
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    dim_config = {
        'dim1': {
            'title': '维度一：员工互评（同级反馈）',
            'columns': ['评估人(EN)', '评估人(中文)', '评估人部门', '被评价同事', '被评价部门', '协作项目',
                       '沟通协作', '专业能力', '责任心', '团队意识', '解决问题', '总均分', '主要优势', '改进建议', '状态', '提交时间'],
            'query': '''SELECT u.en_name, u.ch_name, u.department, d.target_name, d.target_dept,
                d.collaboration_project, d.score_communication, d.score_professional,
                d.score_responsibility, d.score_teamwork, d.score_problem_solving,
                d.strengths, d.improvements, d.submitted, d.updated_at
                FROM dim1_peer d JOIN users u ON d.evaluator_id = u.id ORDER BY u.department, d.id'''
        },
        'dim2': {
            'title': '维度二：员工对直属上级的反馈',
            'columns': ['评估人(EN)', '评估人(中文)', '评估人部门', '上级姓名', '目标制定', '沟通倾听',
                       '授权信任', '反馈指导', '团队氛围', '公平公正', '总均分', '做得好的方面', '需改进方面', '建议', '状态', '提交时间'],
            'query': '''SELECT u.en_name, u.ch_name, u.department, d.manager_name,
                d.score_goal_setting, d.score_communication, d.score_delegation, d.score_feedback,
                d.score_team_climate, d.score_fairness, d.strengths, d.improvements,
                d.suggestions, d.submitted, d.updated_at
                FROM dim2_upward d JOIN users u ON d.evaluator_id = u.id ORDER BY u.department'''
        },
        'dim3': {
            'title': '维度三：直属上级对员工的反馈',
            'columns': ['评估人(EN)', '评估人(中文)', '评估人部门', '下属姓名', '下属岗位', '工作质量',
                       '专业能力', '主动性', '团队协作', '解决问题', '客户导向', '总均分', '核心优势', '改进建议', '状态', '提交时间'],
            'query': '''SELECT u.en_name, u.ch_name, u.department, d.subordinate_name,
                d.subordinate_position, d.score_quality, d.score_professional, d.score_initiative,
                d.score_teamwork, d.score_problem_solving, d.score_customer,
                d.strengths, d.improvements, d.submitted, d.updated_at
                FROM dim3_downward d JOIN users u ON d.evaluator_id = u.id ORDER BY u.department, d.id'''
        },
        'dim4': {
            'title': '维度四：匿名领导力反馈',
            'columns': ['匿名标识', '评价对象', '协作关系', '战略视野', '沟通影响', '赋能团队', '变革创新',
                       '诚信正直', '结果导向', '跨部门协作', '情绪管理', '总均分', '反馈建议',
                       '组织管理建议', '沟通透明度建议', '敬业度建议', '状态', '提交时间'],
            'query': '''SELECT anonymous_token, target_name, relationship, score_strategic,
                score_communication as comm, score_empowerment, score_innovation,
                score_integrity, score_execution, score_collaboration, score_emotional,
                feedback_text, q1_org_mgmt, q2_transparency, q3_engagement,
                submitted, updated_at FROM dim4_leadership ORDER BY id'''
        }
    }

    config = dim_config.get(dim)
    if not config:
        return jsonify({'error': '无效维度'}), 400

    ws.merge_cells('A1:Q1')
    c = ws.cell(row=1, column=1, value=f"Petra品牌中国区 - {config['title']} ({datetime.now().strftime('%Y-%m-%d')})")
    c.font = Font(name='Arial', size=14, bold=True, color='1F4E79')
    c.alignment = Alignment(horizontal='center')

    for i, col_name in enumerate(config['columns'], 1):
        c = ws.cell(row=3, column=i, value=col_name)
        c.font = header_font; c.fill = header_fill; c.border = thin_border
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    rows = db.execute(config['query']).fetchall()
    score_cols = []
    for i, col_name in enumerate(config['columns']):
        if col_name in ['沟通协作','专业能力','责任心','团队意识','解决问题','目标制定','沟通倾听','授权信任','反馈指导','团队氛围','公平公正',
                        '工作质量','主动性','团队协作','解决问题','客户导向','战略视野','沟通影响','赋能团队','变革创新','诚信正直','结果导向','跨部门协作','情绪管理']:
            score_cols.append(i)

    for r, row in enumerate(rows):
        for c, val in enumerate(row):
            cell = ws.cell(row=r+4, column=c+1, value=val if val else '')
            cell.font = Font(name='Arial', size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        row_num = r + 4
        try:
            avg_idx = config['columns'].index('总均分')
            cell = ws.cell(row=row_num, column=avg_idx+1)
            if not cell.value and score_cols:
                cell.value = f'=ROUND(AVERAGEIF({openpyxl.utils.get_column_letter(score_cols[0]+1)}{row_num}:{openpyxl.utils.get_column_letter(score_cols[-1]+1)}{row_num},">0"),1)'
        except ValueError:
            pass

    for i in range(1, len(config['columns'])+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Petra_360_{dim}_{datetime.now().strftime("%Y%m%d")}.xlsx')

def export_csv(db, dim):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Sheet'])

    queries = {
        'dim1': 'SELECT * FROM dim1_peer',
        'dim2': 'SELECT * FROM dim2_upward',
        'dim3': 'SELECT * FROM dim3_downward',
        'dim4': 'SELECT * FROM dim4_leadership'
    }
    rows = db.execute(queries.get(dim, queries['dim1'])).fetchall()
    if rows:
        writer.writerow([k for k in rows[0].keys()])
        for row in rows:
            writer.writerow([v for v in row])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode('utf-8-sig')),
                     mimetype='text/csv', as_attachment=True,
                     download_name=f'Petra_360_{dim}_{datetime.now().strftime("%Y%m%d")}.csv')

# ========== Auto-save API ==========
@app.route('/api/autosave', methods=['POST'])
@require_auth
def autosave():
    data = request.get_json()
    dim = data.get('dim')
    items = data.get('items', [])
    db = get_db()
    now = datetime.now().isoformat()

    table_map = {
        'dim1': ('dim1_peer', ['target_name','target_dept','collaboration_project','score_communication','score_professional','score_responsibility','score_teamwork','score_problem_solving','strengths','improvements']),
        'dim2': ('dim2_upward', ['manager_name','score_goal_setting','score_communication','score_delegation','score_feedback','score_team_climate','score_fairness','strengths','improvements','suggestions']),
        'dim3': ('dim3_downward', ['subordinate_name','subordinate_position','score_quality','score_professional','score_initiative','score_teamwork','score_problem_solving','score_customer','strengths','improvements']),
    }

    if dim not in table_map:
        return jsonify({'error': '无效维度'}), 400

    table, fields = table_map[dim]
    for item in items:
        item_id = item.get('id')
        if not item_id:
            continue  # Skip items without ID (not yet saved via explicit save)
        # Skip submitted items (defense against frontend bypass)
        existing = db.execute(f"SELECT submitted FROM {table} WHERE id=?", (item_id,)).fetchone()
        if existing and existing['submitted'] and not g.user.get('can_edit'):
            continue
        vals = [item.get(f, '') for f in fields]
        set_clause = ', '.join([f'{f}=?' for f in fields])
        db.execute(f"UPDATE {table} SET {set_clause}, updated_at=? WHERE id=? AND evaluator_id=?",
                   vals + [now, item_id, g.user['id']])
    db.commit()
    return jsonify({'message': '自动保存成功', 'time': now})

# ========== Submission Status ==========
@app.route('/api/status', methods=['GET'])
@require_auth
def get_status():
    db = get_db()
    uid = g.user['id']
    status = {
        'dim1': db.execute("SELECT COUNT(*) FROM dim1_peer WHERE evaluator_id=? AND submitted=1", (uid,)).fetchone()[0],
        'dim1_total': db.execute("SELECT COUNT(*) FROM dim1_peer WHERE evaluator_id=?", (uid,)).fetchone()[0],
        'dim2': db.execute("SELECT COUNT(*) FROM dim2_upward WHERE evaluator_id=? AND submitted=1", (uid,)).fetchone()[0],
        'dim3': db.execute("SELECT COUNT(*) FROM dim3_downward WHERE evaluator_id=? AND submitted=1", (uid,)).fetchone()[0],
    }
    token = request.args.get('anon_token', '')
    if token:
        status['dim4'] = db.execute("SELECT COUNT(*) FROM dim4_leadership WHERE anonymous_token=? AND submitted=1", (token,)).fetchone()[0]
    return jsonify(status)

@app.route('/api/overall-status', methods=['GET'])
@require_auth
def get_overall_status():
    """Return overall completion status: dim2 + dim3 both submitted = complete."""
    db = get_db()
    uid = g.user['id']
    dim2_cnt = db.execute(
        "SELECT COUNT(*) as cnt FROM dim2_upward WHERE evaluator_id=? AND submitted=1", (uid,)
    ).fetchone()['cnt']
    dim3_cnt = db.execute(
        "SELECT COUNT(*) as cnt FROM dim3_downward WHERE evaluator_id=? AND submitted=1", (uid,)
    ).fetchone()['cnt']
    dim3_total = db.execute(
        "SELECT COUNT(*) as cnt FROM dim3_downward WHERE evaluator_id=?", (uid,)
    ).fetchone()['cnt']
    dim1_cnt = db.execute(
        "SELECT COUNT(*) as cnt FROM dim1_peer WHERE evaluator_id=? AND submitted=1", (uid,)
    ).fetchone()['cnt']
    dim1_total = db.execute(
        "SELECT COUNT(*) as cnt FROM dim1_peer WHERE evaluator_id=?", (uid,)
    ).fetchone()['cnt']
    token = request.args.get('anon_token', '')
    dim4_cnt = 0
    if token:
        dim4_cnt = db.execute(
            "SELECT COUNT(*) as cnt FROM dim4_leadership WHERE anonymous_token=? AND submitted=1", (token,)
        ).fetchone()['cnt']
    return jsonify({
        'dim2_complete': dim2_cnt > 0,
        'dim3_complete': dim3_cnt > 0,
        'dim3_total': dim3_total,
        'dim1_complete': dim1_cnt > 0,
        'dim1_total': dim1_total,
        'dim4_submitted': dim4_cnt,
        'mandatory_complete': dim2_cnt > 0 and dim3_cnt > 0,
        'overall_complete': dim2_cnt > 0 and dim3_cnt > 0
    })

# ========== Super Admin: Admin Role Management ==========
def _log_admin_action(db, operator, target_id, target_name, action, old_value, new_value, detail=''):
    """Record an admin permission change to the operation log."""
    db.execute(
        "INSERT INTO admin_logs (operator_id, operator_name, target_id, target_name, action, old_value, new_value, detail) VALUES (?,?,?,?,?,?,?,?)",
        (operator['id'], operator['en_name'], target_id, target_name, action, old_value, new_value, detail)
    )
    db.commit()

@app.route('/api/admin/manage-admins', methods=['GET'])
@require_super_admin
def list_manageable_admins():
    """List all users with admin role for super admin management."""
    db = get_db()
    admins = db.execute('''
        SELECT id, en_name, ch_name, department, position, role, admin_level, status, created_at
        FROM users WHERE role='admin' AND status='active'
        ORDER BY CASE admin_level
            WHEN 'super_admin' THEN 0
            WHEN 'admin' THEN 1
            ELSE 2
        END, id
    ''').fetchall()
    return jsonify([dict(a) for a in admins])

@app.route('/api/admin/manage-admins/<int:target_id>', methods=['PUT'])
@require_super_admin
def update_admin_level(target_id):
    """Update an admin's privilege level. Only super_admin can call this."""
    db = get_db()
    operator = g.user

    # Prevent self-demotion
    if target_id == operator['id']:
        return jsonify({'error': '不能修改自己的管理员级别'}), 400

    target = db.execute("SELECT * FROM users WHERE id=? AND status='active'", (target_id,)).fetchone()
    if not target:
        return jsonify({'error': '目标用户不存在或已禁用'}), 404
    if target['role'] != 'admin':
        return jsonify({'error': '目标用户不是管理员'}), 400

    data = request.get_json()
    new_level = data.get('admin_level', '').strip()
    if new_level not in ('super_admin', 'admin', ''):
        return jsonify({'error': '无效的管理员级别。可选: super_admin, admin, 或留空（移除管理员权限）'}), 400

    old_level = target['admin_level'] or 'admin'  # treat empty legacy as admin

    # Prevent super_admin from downgrading another super_admin (only self-demotion blocked above)
    # Actually super_admin CAN manage other super_admins — this is by design

    level_labels = {'super_admin': '超级管理员', 'admin': '普通管理员', '': '移除管理员权限'}
    detail = f"由 {operator['en_name']} 将 {target['en_name']} 的管理员级别从「{level_labels.get(old_level, old_level)}」修改为「{level_labels.get(new_level, new_level)}」"

    db.execute("UPDATE users SET admin_level=? WHERE id=?", (new_level, target_id))

    _log_admin_action(db, operator, target_id, target['en_name'], 'change_admin_level',
                      old_level, new_level, detail)

    return jsonify({
        'message': '管理员权限已更新',
        'target': {'id': target_id, 'en_name': target['en_name'], 'admin_level': new_level},
        'detail': detail
    })

@app.route('/api/admin/operation-logs', methods=['GET'])
@require_super_admin
def get_operation_logs():
    """Get admin permission change logs. Supports pagination."""
    db = get_db()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    offset = (page - 1) * per_page

    total = db.execute("SELECT COUNT(*) FROM admin_logs").fetchone()[0]
    logs = db.execute('''
        SELECT * FROM admin_logs ORDER BY created_at DESC LIMIT ? OFFSET ?
    ''', (per_page, offset)).fetchall()

    return jsonify({
        'logs': [dict(l) for l in logs],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': max(1, (total + per_page - 1) // per_page)
    })

# ========== Run ==========
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
